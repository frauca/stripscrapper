"""Excel Exporter - Exporta classificacions a Excel."""

from datetime import datetime
from typing import List
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from stripscraper.models import GlobalClassification


class ExcelExporter:

    def export(self, classifications: List[GlobalClassification], filepath: Path):
        for classification in classifications:
            self._export_classification(classification, filepath)

    def _export_classification(self, classification: GlobalClassification, filepath: Path):
        file_name = f"{classification.category}-{datetime.today().strftime('%Y-%m-%d')}.xlsx"
        excel_file = Path(filepath / file_name)
        excel_file.unlink(missing_ok=True)
        excel_file.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Exporting {excel_file}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Classificació"

        headers = [
            'Posició Global',
            'Equip',
            'Classificació',
            'Punts',
            'Grup',
            'Posició Grup',
            'Partits',
            'Victòries',
            'Derrotes',
            'Sets Favor',
            'Sets Contra',
            'Dif Sets',
            'Punts Favor',
            'Punts Contra',
            'Dif Punts'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for i, team in enumerate(classification.teams, start=2):
            ws.cell(row=i, column=1, value=i-1)
            ws.cell(row=i, column=2, value=team.stats.name)
            ws.cell(row=i, column=3, value=int(team.stats.points_percentage))
            ws.cell(row=i, column=4, value=team.stats.total_points)
            ws.cell(row=i, column=5, value=team.group)
            ws.cell(row=i, column=6, value=team.stats.position)
            ws.cell(row=i, column=7, value=team.stats.matches_played)
            ws.cell(row=i, column=8, value=team.stats.matches_won)
            ws.cell(row=i, column=9, value=team.stats.matches_lost)
            ws.cell(row=i, column=10, value=team.stats.sets_for)
            ws.cell(row=i, column=11, value=team.stats.sets_against)
            ws.cell(row=i, column=12, value=team.stats.sets_difference)
            ws.cell(row=i, column=13, value=team.stats.points_for)
            ws.cell(row=i, column=14, value=team.stats.points_against)
            ws.cell(row=i, column=15, value=team.stats.points_difference)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        wb.save(excel_file)